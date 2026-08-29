"""
Wave 2 of the Metrc drop: every remaining report, keyed to the licence read from
the FILE, never from the filename.

WHY THAT DISTINCTION IS THE WHOLE POINT OF THIS FILE. The first pass judged licence
by filename and got it wrong in both directions:

  * PackagesAdjustmentsReport.xls, PackagesInventoryReport.xls, TransfersReport.xls,
    Transfers(limited)Report.xls, HarvestsReport.xls and LabResultsReport.csv are all
    MC281714. Every NUMBERED sibling - (1)..(6) - is MP281909. The unnumbered file is
    the cultivation pull; the numbered ones are manufacturing. Reading names, the MC
    adjustments book (763 rows) looked absent.

  * Worse, in the other direction: Metrc-Massachusetts-MC281714-Packages-Transferred
    .xlsx and (1).xlsx are named MC281714 and their rows are MP281909. Trusting the
    name would have written manufacturing transfer lines into the cultivation book.
    They load as MP and carry FILENAME_LICENCE_MISMATCH on their source_export row.

AS-OF VARIANTS ARE NOT DUPLICATES. Plants-Vegetative (1) holds 33 rows - the 6 Aug
state, matching the PIT exactly - and (2) holds 1,080, the state after Flower Room #4
came down on 17 Aug. Same report, two different days. Both are kept, in separate
tables, because collapsing them would destroy the only historical plant state we have.
Flowering (1)=4,380 and (2)=3,330 are the same pair and kept the same way.

LAB (1) AND (2) ARE THE SAME SIZE AND DIFFERENT BYTES - sha256 b3000abd... vs
c4177cab... - so the load-once rule does not apply. Both load.

Inventory reports carry SNAPSHOT_NOT_ON_HAND. On-hand is Packages-Active and nothing
else. No PIT table is touched and no close is loaded.
"""
import csv, hashlib, os, re, datetime, warnings
warnings.filterwarnings("ignore")

DOWNLOADS = r"C:\Users\demar\Downloads"
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "supabase", "migrations")

MISMATCH = "FILENAME_LICENCE_MISMATCH: filename says MC281714, rows say MP281909. Loaded as MP281909."
SNAP = "SNAPSHOT_NOT_ON_HAND: a windowed Packages Inventory export. Never on-hand, never Packages Active."

# file, licence FROM CONTENT, table, note
SPEC = [
    # ── MC281714, the unnumbered family ──────────────────────────────────────
    ("PackagesAdjustmentsReport.xls",            "MC281714", "stg_mc_package_adjustments", ""),
    ("PackagesInventoryReport.xls",              "MC281714", "stg_mc_inventory_snapshot",  SNAP),
    ("TransfersReport.xls",                      "MC281714", "stg_mc_transfers_report",    ""),
    ("Transfers(limited)Report.xls",             "MC281714", "stg_mc_transfers_limited",   ""),
    ("HarvestsReport (1).xls",                   "MC281714", "stg_mc_harvests_report",     ""),
    ("PlantsDestroyedReport (1).xls",            "MC281714", "stg_mc_plants_destroyed",    ""),
    ("PlantsInventoryReport.xls",                "MC281714", "stg_mc_plants_inventory",    ""),
    ("PlantsTrendReport (2).xls",                "MC281714", "stg_mc_plants_trend",        ""),
    # ── MC281714 grids and catalogs ──────────────────────────────────────────
    ("Metrc-Massachusetts-MC281714-Packages-Active (1).xlsx",   "MC281714", "stg_mc_packages_active",   ""),
    ("Metrc-Massachusetts-MC281714-Packages-Inactive (2).xlsx", "MC281714", "stg_mc_packages_inactive", ""),
    ("Metrc-Massachusetts-MC281714-Items (1).xlsx",             "MC281714", "stg_mc_items",             ""),
    ("Metrc-Massachusetts-MC281714-Strains (1).xlsx",           "MC281714", "stg_mc_strains",           ""),
    ("Metrc-Massachusetts-MC281714-Locations (2).xlsx",         "MC281714", "stg_mc_locations",         ""),
    ("Metrc-Massachusetts-MC281714-Employees.xlsx",             "MC281714", "stg_mc_employees",         ""),
    ("Metrc-Massachusetts-MC281714-TagOrders-History (1).xlsx", "MC281714", "stg_mc_tag_orders",        ""),
    ("Metrc-Massachusetts-MC281714-LicensedTransfers-Outgoing (1).xlsx",        "MC281714", "stg_mc_transfers_outgoing", ""),
    ("Metrc-Massachusetts-MC281714-LicensedTransfers-Incoming (1).xlsx",        "MC281714", "stg_mc_transfers_incoming", ""),
    ("Metrc-Massachusetts-MC281714-LicensedTransfers-IncomingInactive (1).xlsx","MC281714", "stg_mc_transfers_incoming_inactive", ""),
    # ── as-of variants, deliberately separate tables ─────────────────────────
    ("Metrc-Massachusetts-MC281714-Plants-Vegetative (1).xlsx", "MC281714", "stg_mc_plants_vegetative_asof_0806",
     "AS-OF 6 Aug 2026 state, 33 rows, matches the PIT. NOT a duplicate of the 1,080-row live pull."),
    ("Metrc-Massachusetts-MC281714-Plants-Flowering (1).xlsx",  "MC281714", "stg_mc_plants_flowering_asof_0806",
     "AS-OF 6 Aug 2026 state, 4,380 rows, matches the PIT. NOT a duplicate of the 3,330-row live pull."),
    # ── MP281909 ─────────────────────────────────────────────────────────────
    ("Metrc-Massachusetts-MP281909-Locations.xlsx",                             "MP281909", "stg_mp_locations", ""),
    ("Metrc-Massachusetts-MP281909-LicensedTransfers-Outgoing (4).xlsx",        "MP281909", "stg_mp_transfers_outgoing", ""),
    ("Metrc-Massachusetts-MP281909-LicensedTransfers-IncomingInactive (2).xlsx","MP281909", "stg_mp_transfers_incoming_inactive", ""),
    ("TransfersReport.csv",                      "MP281909", "stg_mp_transfers_report",    ""),
    ("WholesaleTransfersReport (7).xls",         "MP281909", "stg_mp_wholesale_transfers", ""),
    # ── the two mislabelled files ────────────────────────────────────────────
    ("Metrc-Massachusetts-MC281714-Packages-Transferred.xlsx",     "MP281909", "stg_mp_packages_transferred_altpull", MISMATCH),
    ("Metrc-Massachusetts-MC281714-Packages-Transferred (1).xlsx", "MP281909", "stg_mp_packages_transferred_altpull", MISMATCH),
    # ── lab ──────────────────────────────────────────────────────────────────
    ("LabResultsReport.csv",   "MC281714", "stg_mc_lab_results", "The MC lab book."),
    ("LabResultsReport.xls",   "MP281909", "stg_mp_lab_results", ""),
    ("LabResultsReport (1).xls","MP281909","stg_mp_lab_results", "sha256 b3000abd... - differs from (2), not a duplicate."),
    ("LabResultsReport (2).xls","MP281909","stg_mp_lab_results", "sha256 c4177cab... - differs from (1), not a duplicate."),
    ("LabResultsReport (3).xls","MP281909","stg_mp_lab_results", ""),
]


def slug(s, used):
    s = re.sub(r"[^a-z0-9]+", "_", (s or "").strip().lower()).strip("_") or "col"
    if s[0].isdigit():
        s = "c_" + s
    base, n = s, 2
    while s in used:
        s, n = f"{base}_{n}", n + 1
    used.add(s)
    return s


def lit(v):
    if v is None:
        return "null"
    s = str(v)
    return "null" if s == "" else "'" + s.replace("'", "''") + "'"


def read_any(path):
    e = os.path.splitext(path)[1].lower()
    if e == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        g = [["" if c.value is None else str(c.value).strip() for c in r] for r in ws.iter_rows()]
        wb.close()
        return g, ""
    if e == ".csv":
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            return [r for r in csv.reader(fh)], ""
    import xlrd
    b = xlrd.open_workbook(path)
    s = b.sheet_by_index(0)
    g = [[str(s.cell_value(r, c)).strip() for c in range(s.ncols)] for r in range(s.nrows)]
    win, hdr = "", 0
    for r, row in enumerate(g[:25]):
        for v in row:
            if v.startswith("From "):
                win = v
        if sum(1 for v in row if v) >= 4 and any(
                v in ("Package", "Tag", "Package Tag", "Type", "Tag Number",
                      "Destination License", "Harvest", "Plant", "Manifest") for v in row):
            hdr = r
            break
    return g[hdr:], win


def main():
    stamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    tables, se, report = {}, [], []
    for fname, lic, table, note in SPEC:
        p = os.path.join(DOWNLOADS, fname)
        if not os.path.exists(p):
            report.append((fname, lic, "MISSING", 0, 0, table)); continue
        raw = open(p, "rb").read()
        sha = hashlib.sha256(raw).hexdigest()
        grid, win = read_any(p)
        if not grid:
            report.append((fname, lic, "EMPTY", 0, 0, table)); continue
        used = set()
        header = [slug(h, used) for h in grid[0]]
        body = [r for r in grid[1:] if any((c or "").strip() for c in r)]
        t = tables.setdefault(table, {"cols": [], "rows": []})
        for c in header:
            if c not in t["cols"]:
                t["cols"].append(c)
        for r in body:
            t["rows"].append((dict(zip(header, r)), fname, sha, lic, win))
        se.append((fname, lic, win, len(body), sha,
                   datetime.date.fromtimestamp(os.path.getmtime(p)), note))
        report.append((fname, lic, "OK", len(body), 0, table))

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    p0 = os.path.join(OUT_DIR, f"{stamp}_ingest_w2_00_source_export.sql")
    with open(p0, "w", encoding="utf-8", newline="\n") as f:
        f.write(__doc__.replace("\n", "\n-- ").join(["-- ", "\n"]))
        for fn, lic, win, n, sha, mt, note in se:
            proof = (note + " ") if note else ""
            f.write("insert into public.source_export(file_name,system,report,licence,period_stated,"
                    "period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                    f"{lit(fn)},'metrc',{lit(fn.split('.')[0])},{lit(lic)},{lit(win)},{lit(win)},{n},"
                    f"{lit(sha[:16])},false,{lit(proof + 'Licence read from the file, not the filename.')},"
                    f"{lit(str(mt))}) on conflict do nothing;\n")
    written.append(p0)

    # NO SINGLE MIGRATION MAY GET LARGE. The MP lab book alone came to 86.8 MB in one
    # file - close to GitHub's hard ceiling, and git history is permanent, so an
    # oversized file is a mistake you cannot take back. Tables are split into parts of
    # at most PART_ROWS, each part independently appliable; the CREATE lives in part 1.
    PART_ROWS = 12000
    i = 0
    for table, t in sorted(tables.items()):
        cols = t["cols"]
        allc = cols + ["source_file", "file_sha256", "licence", "file_window"]
        parts = [t["rows"][k:k + PART_ROWS] for k in range(0, len(t["rows"]), PART_ROWS)] or [[]]
        for pn, rows in enumerate(parts, start=1):
            i += 1
            suffix = table if len(parts) == 1 else f"{table}_p{pn}of{len(parts)}"
            p = os.path.join(OUT_DIR, f"{stamp}_ingest_w2_{i:02d}_{suffix}.sql")
            with open(p, "w", encoding="utf-8", newline="\n") as f:
                f.write(f"-- {table}: part {pn} of {len(parts)}, {len(rows)} rows. NOT APPLIED, branch only.\n"
                        "-- Licence on every row comes from the FILE, never the filename.\n")
                if pn == 1:
                    f.write(f"create table if not exists public.{table} (\n")
                    f.write(",\n".join(f"  {c} text" for c in cols))
                    f.write(",\n  source_file text not null,\n  file_sha256 text not null,\n  licence text not null,\n"
                            "  file_window text,\n  ingested_at timestamptz not null default now()\n);\n")
                    f.write(f"alter table public.{table} enable row level security;\n"
                            f"drop policy if exists {table}_read on public.{table};\n"
                            f"create policy {table}_read on public.{table} for select to authenticated using (true);\n")
                for j in range(0, len(rows), 500):
                    ch = rows[j:j + 500]
                    f.write(f"insert into public.{table} ({','.join(allc)}) values\n")
                    f.write(",\n".join("(" + ",".join([lit(d.get(c)) for c in cols] +
                            [lit(fn), lit(sha), lit(lc), lit(w)]) + ")" for d, fn, sha, lc, w in ch))
                    f.write(";\n")
            written.append(p)

    print(f"{'file':56}|{'licence':9}|{'status':8}|{'rows':>7}|{'rej':>4}| table")
    for fn, lic, st, n, rej, tb in report:
        print(f"{fn[:54]:56}|{lic:9}|{st:8}|{n:>7}|{rej:>4}| {tb}")
    print(f"\n{len(written)} files, {sum(os.path.getsize(x) for x in written):,} bytes, "
          f"{sum(len(t['rows']) for t in tables.values()):,} rows, {len(tables)} tables")


if __name__ == "__main__":
    main()
