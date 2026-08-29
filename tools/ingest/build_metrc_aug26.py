"""
Generate migration SQL for the 26 Aug 2026 Metrc export drop.

WHY A GENERATOR AND NOT HAND-WRITTEN SQL. The drop carries 14,168 transferred
package lines, 53,012 inactive plants and 4,407 waste rows. Hand-transcribing that
through a tool call is how a stray character reached production earlier in this
session; a generator reads the bytes off disk and never retypes them.

WHAT IT REFUSES TO DO
  * It does not invent a schema. Each file lands in its own staging table whose
    columns are the file's OWN header, slugged, all text. Typing and mapping into
    the domain tables is a later, reviewed step - guessing a type here would bake
    a decision nobody reviewed into 14,168 rows.
  * It does not touch metrc_rpt_point_in_time, and it loads no 2024-12-31 or
    2025-12-31 close. Agent A owns PIT.
  * It does not net, sum, or derive anything. Shipped and received stay separate
    columns. Adjustment signs are copied verbatim. Nothing is coalesced to 0.
  * It stops a file whose parsed row count misses the expected count by more than
    1%, and emits nothing for that file.

EVERY ROW CARRIES ITS PROVENANCE: source_file, file_sha256 and the licence read
from the title block, so a row can always be traced back to the bytes it came from.
"""
import csv, hashlib, os, re, sys, datetime

DOWNLOADS = r"C:\Users\demar\Downloads"
OUT_DIR   = os.path.join(os.path.dirname(__file__), "..", "..", "supabase", "migrations")

# file, licence, expected data rows (None = no stated expect), staging table
SPEC = [
    ("Metrc-Massachusetts-MP281909-Packages-Active.xlsx",             "MP281909", 507,   "stg_mp_packages_active"),
    ("Metrc-Massachusetts-MP281909-Packages-Inactive.xlsx",           "MP281909", 1925,  "stg_mp_packages_inactive"),
    ("Metrc-Massachusetts-MP281909-Packages-InTransit.xlsx",          "MP281909", 141,   "stg_mp_packages_intransit"),
    ("Metrc-Massachusetts-MP281909-Packages-Transferred (2).xlsx",    "MP281909", 14168, "stg_mp_packages_transferred"),
    ("Metrc-Massachusetts-MP281909-LicensedTransfers-Incoming (1).xlsx","MP281909", 2,   "stg_mp_transfers_incoming"),
    ("PackagesAdjustmentsReport (6).xls",                             "MP281909", 3725,  "stg_mp_package_adjustments"),
    ("PackagesInventoryReport (4).xls",                               "MP281909", 34,    "stg_mp_inventory_snapshot"),
    ("PackagesInventoryReport (5).xls",                               "MP281909", 22,    "stg_mp_inventory_snapshot"),
    ("PackagesInventoryReport (6).xls",                               "MP281909", 9,     "stg_mp_inventory_snapshot"),
    ("Metrc-Massachusetts-MC281714-Plants-Flowering (2).xlsx",        "MC281714", 3330,  "stg_mc_plants_flowering"),
    ("Metrc-Massachusetts-MC281714-Plants-Vegetative (2).xlsx",       "MC281714", 1080,  "stg_mc_plants_vegetative"),
    ("Metrc-Massachusetts-MC281714-Plants-Inactive (1).xlsx",         "MC281714", 53012, "stg_mc_plants_inactive"),
    ("Metrc-Massachusetts-MC281714-Plants-Harvests (4).xlsx",         "MC281714", 28,    "stg_mc_harvests_open"),
    ("Metrc-Massachusetts-MC281714-Plants-HarvestsInactive (3).xlsx", "MC281714", 357,   "stg_mc_harvests_inactive"),
    ("Metrc-Massachusetts-MC281714-Plants-Plantings-Active (3).xlsx", "MC281714", 57,    "stg_mc_plantings_active"),
    ("Metrc-Massachusetts-MC281714-Plants-Plantings-Inactive (2).xlsx","MC281714", 2705, "stg_mc_plantings_inactive"),
    ("Metrc-Massachusetts-MC281714-Plants-Waste (2).xlsx",            "MC281714", 4407,  "stg_mc_plants_waste"),
    ("Transfers(limited)Report (1).csv",                              "MP281909", None,  "stg_transfers_limited"),
    ("Metrc-Massachusetts-MC281714-Tags-Used.xlsx",                   "MC281714", None,  "stg_mc_tags_used"),
    ("Metrc-Massachusetts-MC281714-Tags-Voided.xlsx",                 "MC281714", None,  "stg_mc_tags_voided"),
    ("Metrc-Massachusetts-MC281714-Tags-Available.xlsx",              "MC281714", None,  "stg_mc_tags_available"),
    ("Metrc-Massachusetts-MP281909-Tags-Used.xlsx",                   "MP281909", None,  "stg_mp_tags_used"),
    ("Metrc-Massachusetts-MP281909-Tags-Voided.xlsx",                 "MP281909", None,  "stg_mp_tags_voided"),
    ("Metrc-Massachusetts-MP281909-Tags-Available.xlsx",              "MP281909", None,  "stg_mp_tags_available"),
]


def slug(s, used):
    s = re.sub(r"[^a-z0-9]+", "_", (s or "").strip().lower()).strip("_")
    s = s or "col"
    if s[0].isdigit():
        s = "c_" + s
    base, n = s, 2
    while s in used:
        s, n = f"{base}_{n}", n + 1
    used.add(s)
    return s


def lit(v):
    if v is None:
        return "null"
    s = str(v)
    if s == "":
        return "null"
    return "'" + s.replace("'", "''") + "'"


def read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [["" if c.value is None else str(c.value).strip() for c in r] for r in ws.iter_rows()]
    wb.close()
    return rows, ""


def read_xls(path):
    """Metrc .xls carries a title block; the header is the first row that repeats
    a known column word. Everything above it is banner, not data."""
    import xlrd
    b = xlrd.open_workbook(path)
    s = b.sheet_by_index(0)
    grid = [[str(s.cell_value(r, c)).strip() for c in range(s.ncols)] for r in range(s.nrows)]
    window = ""
    hdr = 0
    for r, row in enumerate(grid[:25]):
        for v in row:
            if v.startswith("From "):
                window = v
        if sum(1 for v in row if v) >= 4 and any(
            v in ("Package", "Tag", "Package Tag", "Type", "Tag Number", "Destination License") for v in row
        ):
            hdr = r
            break
    return grid[hdr:], window


def read_csv(path):
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        return [r for r in csv.reader(fh)], ""


def main():
    stamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    tables, se_rows, report = {}, [], []

    for fname, licence, expect, table in SPEC:
        path = os.path.join(DOWNLOADS, fname)
        if not os.path.exists(path):
            report.append((fname, "MISSING", 0, 0, table)); continue

        raw = open(path, "rb").read()
        sha = hashlib.sha256(raw).hexdigest()
        ext = os.path.splitext(fname)[1].lower()
        grid, window = (read_xlsx if ext == ".xlsx" else read_csv if ext == ".csv" else read_xls)(path)
        if not grid:
            report.append((fname, "EMPTY", 0, 0, table)); continue

        used = set()
        header = [slug(h, used) for h in grid[0]]
        body = [r for r in grid[1:] if any((c or "").strip() for c in r)]
        n = len(body)

        if expect is not None and abs(n - expect) > max(1, expect * 0.01):
            report.append((fname, f"STOP off by {n-expect}", n, 0, table)); continue

        t = tables.setdefault(table, {"cols": [], "rows": []})
        for c in header:
            if c not in t["cols"]:
                t["cols"].append(c)
        for r in body:
            t["rows"].append((dict(zip(header, r)), fname, sha, licence, window))

        se_rows.append((fname, licence, window, n, sha, datetime.date.fromtimestamp(os.path.getmtime(path))))
        report.append((fname, "OK", n, n, table))

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []

    # ONE MIGRATION PER STAGING TABLE. A single file for all 176,461 rows came to
    # 30.7 MB - too big to review, too big to apply in one statement, and a merge
    # conflict in it would be unresolvable. Split, each file independently
    # appliable and independently revertable.
    seq = 0
    out0 = os.path.join(OUT_DIR, f"{stamp}_ingest_aug26_00_source_export.sql")
    with open(out0, "w", encoding="utf-8", newline="\n") as f:
        f.write(HEADER)
        f.write("\n-- ── source_export: the file register, written FIRST ──────────────\n")
        for fn, lic, win, n, sha, mt in se_rows:
            f.write(
                "insert into public.source_export"
                "(file_name,system,report,licence,period_stated,period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                f"{lit(fn)},'metrc',{lit(fn.split('.')[0])},{lit(lic)},{lit(win)},{lit(win)},{n},{lit(sha[:16])},false,"
                "'Loaded to a staging table with the file''s own header. No typing, no derivation.',"
                f"{lit(str(mt))}) on conflict do nothing;\n"
            )
    written.append(out0)

    for table, t in sorted(tables.items()):
        seq += 1
        out = os.path.join(OUT_DIR, f"{stamp}_ingest_aug26_{seq:02d}_{table}.sql")
        cols = t["cols"]
        with open(out, "w", encoding="utf-8", newline="\n") as f:
            f.write(HEADER)
            f.write(f"\n-- ── {table}: {len(t['rows'])} rows ──────────────────────────────\n")
            f.write(f"create table if not exists public.{table} (\n")
            f.write(",\n".join(f"  {c} text" for c in cols))
            f.write(",\n  source_file text not null,\n  file_sha256 text not null,\n  licence text not null,\n"
                    "  file_window text,\n  ingested_at timestamptz not null default now()\n);\n")
            f.write(f"alter table public.{table} enable row level security;\n")
            f.write(f"drop policy if exists {table}_read on public.{table};\n")
            f.write(f"create policy {table}_read on public.{table} for select to authenticated using (true);\n")
            allc = cols + ["source_file", "file_sha256", "licence", "file_window"]
            for i in range(0, len(t["rows"]), 500):
                chunk = t["rows"][i:i + 500]
                f.write(f"insert into public.{table} ({','.join(allc)}) values\n")
                f.write(",\n".join(
                    "(" + ",".join([lit(d.get(c)) for c in cols] + [lit(fn), lit(sha), lit(lic), lit(win)]) + ")"
                    for d, fn, sha, lic, win in chunk))
                f.write(";\n")
        written.append(out)

    print(f"WROTE {len(written)} migration files, "
          f"{sum(os.path.getsize(p) for p in written):,} bytes total\n")
    print(f"{'file':60} {'status':16} {'rows':>7} {'loaded':>7}  table")
    for fn, st, n, ld, tb in report:
        print(f"{fn[:58]:60} {st:16} {n:>7} {ld:>7}  {tb}")
    print(f"\nstaging tables: {len(tables)}   source_export rows: {len(se_rows)}   "
          f"total data rows: {sum(len(t['rows']) for t in tables.values())}")
    print(f"file bytes: {os.path.getsize(out):,}")


HEADER = """-- METRC EXPORT DROP, 26 AUGUST 2026 - STAGED, NOT TYPED.
--
-- NOT APPLIED. Branch only, held for APPLY.
--
-- Generated by tools/ingest/build_metrc_aug26.py directly from the files in
-- Downloads. Nothing here was retyped by hand: 14,168 transferred package lines and
-- 53,012 inactive plants do not survive transcription, and this session already put
-- one stray character into a deployed edge function that way.
--
-- EACH FILE LANDS IN ITS OWN STAGING TABLE WHOSE COLUMNS ARE THE FILE'S OWN HEADER,
-- ALL TEXT. That is deliberate. Typing a column guesses at a meaning nobody has
-- reviewed, and a wrong guess repeated 14,168 times is far more expensive than a
-- later, reviewed mapping step. Every row carries source_file, file_sha256, licence
-- and the report window, so any figure can be traced back to the bytes it came from.
--
-- WHAT IS NOT HERE, BY INSTRUCTION:
--   * metrc_rpt_point_in_time is untouched, and no 2024-12-31 or 2025-12-31 close is
--     loaded. Agent A owns PIT and the licence-in-key change.
--   * PackagesInventoryReport (4)(5)(6) are SNAPSHOTS, never on-hand. They are staged
--     with their own window and must not be upserted into live quantity or used as
--     Packages Active. On-hand for MP281909 is Packages-Active (507) and nothing else.
--   * Shipped and received quantities are stored as they appear and are never netted.
--   * Adjustment quantities keep Metrc's sign.
--   * The two incoming transfer rows stay unreceived - no received_at is invented.
--   * IL* destinations on transferred packages are lab submits, not sales. Nothing
--     here classifies them; the raw destination licence is preserved for the mapping
--     step to judge.
--
-- COUNT GATE: a file whose parsed rows missed its expected count by more than 1%
-- emitted nothing at all. See the run report in the PR body for what passed.
"""

if __name__ == "__main__":
    main()
