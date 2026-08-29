"""
Wave 3: the four files dropped 29 Aug, plus the three grids proven EMPTY.

TWO OF THE FIVE "ABSENT" REPORTS WERE NEVER MISSING - THEY ARE EMPTY.
The owner supplied Metrc screenshots showing:

    Packages -> In Transit, MC281714      "No data is available to display."  0 rows
    Transfers -> Rejected, MP281909       "No data is available to display."  0 rows
    Transfers -> Rejected, MC281714       "No data is available to display."  0 rows

An empty grid and an unpulled report are different facts and only one of them is a
gap. These three are recorded in source_export with rows_in_file = 0 and the screen
as the evidence, so nobody re-requests them and nobody reads the silence as a hole.
They have no file, so they carry no hash - stated plainly rather than faked.

THE LICENCE ON THESE FOUR COMES FROM THE FILENAME, AND THAT IS SAID OUT LOUD.
Wave 2 established that the title block beats the filename, because two files named
MC281714 held MP281909 rows. These four are Metrc GRID exports, not Control Panel
reports: they carry no title block and no licence column anywhere in the sheet -
Items has 'Item' in A1, Strains has 'Strain'. So the filename is the only signal
that exists. That is weaker evidence than wave 2 had, and the source_export row says
so, rather than implying a verification that did not happen.

MP281909 Items (712) and Strains (107) close two of the five gaps. Production
Management remains the only report still genuinely absent.
"""
import hashlib, os, re, datetime, warnings
warnings.filterwarnings("ignore")

DOWNLOADS = r"C:\Users\demar\Downloads"
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "supabase", "migrations")
FROM_NAME = ("Licence from the FILENAME. This is a Metrc grid export: it carries no title block and no "
             "licence column, so no in-file confirmation is possible. Weaker evidence than the Control "
             "Panel reports, and recorded as such.")

SPEC = [
    ("Metrc-Massachusetts-MP281909-Packages-InTransit (1).xlsx", "MP281909", "stg_mp_packages_intransit",
     "29 Aug pull, 160 rows, into the same table as the 26 Aug pull of 141. Different days, both kept - "
     "source_file and file_sha256 separate them, and in-transit is a moving position, not a duplicate."),
    ("Metrc-Massachusetts-MP281909-Items.xlsx",                  "MP281909", "stg_mp_items",
     "Closes a gap: MP281909 Items did not exist in the 26 Aug drop."),
    ("Metrc-Massachusetts-MP281909-Strains.xlsx",                "MP281909", "stg_mp_strains",
     "Closes a gap: MP281909 Strains did not exist in the 26 Aug drop."),
]

# Registered but NOT inserted. Its sha256 differs from MC-Items (1) - c6a4ceec vs
# 204ca296 - so a hash test alone would have loaded 494 rows twice. Compared row by
# row instead: same header, same 494 rows, zero rows and zero item keys on either
# side. The byte difference is xlsx internal metadata, not catalog data. Hashes prove
# files differ; only content proves data differs.
DUPLICATE = [
    ("Metrc-Massachusetts-MC281714-Items (2).xlsx", "MC281714",
     "DUPLICATE of Metrc-Massachusetts-MC281714-Items (1).xlsx. sha256 differs (xlsx metadata) but the "
     "catalog is identical: same header, same 494 rows, 0 rows and 0 item keys unique to either side. "
     "Registered and deliberately NOT inserted - loading it would double the MC item catalog."),
]

# grid, licence, evidence - no file, so no hash, and rows_in_file = 0
EMPTY = [
    ("Packages / In Transit",   "MC281714", "Owner screenshot 29 Aug 2026: 'No data is available to display.', Page 0 of 0."),
    ("Transfers / Rejected",    "MP281909", "Owner screenshot 29 Aug 2026: 'No data is available to display.', Page 0 of 0."),
    ("Transfers / Rejected",    "MC281714", "Owner screenshot 29 Aug 2026: 'No data is available to display.', Page 0 of 0."),
]

# Closed, not outstanding. Owner confirmed 29 Aug 2026 the report is not on this
# Metrc tenant at all - so it is neither pulled nor pullable, and nobody should keep
# chasing it. It appeared on the owner's own 29 Aug request list; it was never found
# and was never invented.
UNAVAILABLE = [
    ("Production Management", "MC281714",
     "UNAVAILABLE_ON_TENANT. Owner confirmed 29 Aug 2026 that this report does not exist on this Metrc "
     "tenant. Not a pull gap and not a load failure - there is nothing to export. Closed."),
]


def slug(s, used):
    s = re.sub(r"[^a-z0-9]+", "_", (s or "").strip().lower()).strip("_") or "col"
    if s[0].isdigit():
        s = "c_" + s
    b, n = s, 2
    while s in used:
        s, n = f"{b}_{n}", n + 1
    used.add(s)
    return s


def lit(v):
    if v is None:
        return "null"
    s = str(v)
    return "null" if s == "" else "'" + s.replace("'", "''") + "'"


def main():
    import openpyxl
    stamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    os.makedirs(OUT_DIR, exist_ok=True)
    report, written = [], []

    p0 = os.path.join(OUT_DIR, f"{stamp}_ingest_w3_00_source_export.sql")
    with open(p0, "w", encoding="utf-8", newline="\n") as f:
        f.write("-- Wave 3 file register. NOT APPLIED, branch only.\n")
        for name, lic, why in EMPTY:
            f.write("insert into public.source_export(file_name,system,report,licence,period_stated,"
                    "period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                    f"{lit('(no file - empty grid) ' + name + ' ' + lic)},'metrc',{lit(name)},{lit(lic)},"
                    f"'as at 2026-08-29','as at 2026-08-29',0,null,false,"
                    f"{lit('PROVEN EMPTY, not missing. ' + why + ' No file exists, so there is no hash; an empty grid and an unpulled report are different facts.')},"
                    "'2026-08-29') on conflict do nothing;\n")

        for name, lic, why in UNAVAILABLE:
            f.write("insert into public.source_export(file_name,system,report,licence,period_stated,"
                    "period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                    f"{lit('(no report on tenant) ' + name + ' ' + lic)},'metrc',{lit(name)},{lit(lic)},"
                    f"'n/a','n/a',0,null,false,{lit(why)},'2026-08-29') on conflict do nothing;\n")

        for fname, lic, why in DUPLICATE:
            p = os.path.join(DOWNLOADS, fname)
            if not os.path.exists(p):
                continue
            sha = hashlib.sha256(open(p, "rb").read()).hexdigest()
            f.write("insert into public.source_export(file_name,system,report,licence,period_stated,"
                    "period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                    f"{lit(fname)},'metrc',{lit(fname.split('.')[0])},{lit(lic)},'grid export, no window',"
                    f"'grid export, no window',494,{lit(sha[:16])},false,{lit(why)},"
                    f"{lit(str(datetime.date.fromtimestamp(os.path.getmtime(p))))}) on conflict do nothing;\n")
            report.append((fname, lic, "DUPLICATE", 494, "(not inserted)"))

        for fname, lic, table, note in SPEC:
            p = os.path.join(DOWNLOADS, fname)
            if not os.path.exists(p):
                report.append((fname, lic, "MISSING", 0, table)); continue
            sha = hashlib.sha256(open(p, "rb").read()).hexdigest()
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            n = ws.max_row - 1
            wb.close()
            proof = ((note + " ") if note else "") + FROM_NAME
            f.write("insert into public.source_export(file_name,system,report,licence,period_stated,"
                    "period_actual,rows_in_file,sha256_16,used_in_audit,what_it_proved,captured_on) values("
                    f"{lit(fname)},'metrc',{lit(fname.split('.')[0])},{lit(lic)},'grid export, no window',"
                    f"'grid export, no window',{n},{lit(sha[:16])},false,{lit(proof)},"
                    f"{lit(str(datetime.date.fromtimestamp(os.path.getmtime(p))))}) on conflict do nothing;\n")
            report.append((fname, lic, "OK", n, table))
    written.append(p0)

    for i, (fname, lic, table, note) in enumerate(SPEC, start=1):
        p = os.path.join(DOWNLOADS, fname)
        if not os.path.exists(p):
            continue
        sha = hashlib.sha256(open(p, "rb").read()).hexdigest()
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid = [["" if c.value is None else str(c.value).strip() for c in r] for r in ws.iter_rows()]
        wb.close()
        used = set()
        cols = [slug(h, used) for h in grid[0]]
        body = [r for r in grid[1:] if any((c or "").strip() for c in r)]
        out = os.path.join(OUT_DIR, f"{stamp}_ingest_w3_{i:02d}_{table}.sql")
        allc = cols + ["source_file", "file_sha256", "licence", "licence_evidence"]
        with open(out, "w", encoding="utf-8", newline="\n") as f:
            f.write(f"-- {table}: {len(body)} rows. NOT APPLIED, branch only.\n"
                    "-- Licence from the FILENAME - this grid export carries no title block and no\n"
                    "-- licence column, so no in-file confirmation is possible. Stated, not implied.\n")
            f.write(f"create table if not exists public.{table} (\n")
            f.write(",\n".join(f"  {c} text" for c in cols))
            f.write(",\n  source_file text not null,\n  file_sha256 text not null,\n  licence text not null,\n"
                    "  licence_evidence text not null,\n  ingested_at timestamptz not null default now()\n);\n")
            f.write(f"alter table public.{table} enable row level security;\n"
                    f"drop policy if exists {table}_read on public.{table};\n"
                    f"create policy {table}_read on public.{table} for select to authenticated using (true);\n")
            for j in range(0, len(body), 500):
                ch = body[j:j + 500]
                f.write(f"insert into public.{table} ({','.join(allc)}) values\n")
                f.write(",\n".join("(" + ",".join([lit(dict(zip(cols, r)).get(c)) for c in cols] +
                        [lit(fname), lit(sha), lit(lic), lit('filename only - grid export has no licence field')]) + ")"
                        for r in ch))
                f.write(";\n")
        written.append(out)

    print(f"{'file':58}|{'licence':9}|{'status':8}|{'rows':>6}| table")
    for fn, lic, st, n, tb in report:
        print(f"{fn[:56]:58}|{lic:9}|{st:8}|{n:>6}| {tb}")
    print("\nPROVEN EMPTY (no file, recorded so nobody re-requests them):")
    for name, lic, why in EMPTY:
        print(f"  {name:26} {lic}  0 rows")
    print(f"\n{len(written)} files, {sum(os.path.getsize(x) for x in written):,} bytes")


if __name__ == "__main__":
    main()
