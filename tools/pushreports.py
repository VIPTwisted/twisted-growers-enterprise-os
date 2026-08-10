"""Push the Metrc report exports through the platform's own import pipeline.

Detection and mapping happen in the database (tg_import_report_do) - this only
turns a spreadsheet into rows and establishes the as-of date from inside the
file, never from when it was uploaded.

NOTE on reading .xlsx: openpyxl's read_only mode trusts the <dimension> record,
and Metrc writes that as a single column. It is wrong. Always load with
read_only=False, or every export appears to have one column.
"""
import glob, json, os, re, sys, time, urllib.error, urllib.request, warnings
import openpyxl, xlrd

warnings.filterwarnings('ignore')

FN = 'https://fxetuqjryttnypgepsru.supabase.co/functions/v1/report-ingest'
HDR = {'x-admin-key': os.environ.get('TG_ADMIN_KEY', ''),
       'Authorization': 'Bearer ' + os.environ.get('SUPA_KEY', ''),
       'Content-Type': 'application/json'}

OUR_LICENCES = {'MC281714', 'MP281909'}

DATE_FIELDS = ['Finished', 'Date', 'Received Date', 'Adj. Date', 'Waste Date',
               'Phase Date', 'Created', 'Received', 'Test Date', 'Finished Date',
               'Batch Date', 'Pack Date']


def grid_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)   # never read_only
    sh = wb[wb.sheetnames[0]]
    return [[('' if c.value is None else str(c.value)).strip() for c in row]
            for row in sh.iter_rows()]


def grid_xls(path):
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    return [[str(sh.cell_value(r, c)).strip() for c in range(sh.ncols)]
            for r in range(sh.nrows)]


SUBDETAIL_LABELS = {'destroyed note', 'destroyed date', 'destroyed reason', 'user',
                    'waste note', 'waste date', 'adjustment note'}

FOOTER = re.compile(r'\s*(totals?|grand total|subtotal|sum)\s*:?\s*$', re.I)


def to_rows(grid, carry=None):
    """Read a Metrc RENDERED REPORT, not a spreadsheet.

    These files are banded output - banner, column header, optional group header,
    detail rows, optional sub-detail rows, group and grand-total footers. Treating
    them as flat tables is what caused every import fault in this session:
    a footer became a $1.69M sale, 396 grouped rows were dropped, and every plant
    was counted twice because its destruction detail sits on a second row.

    Three bands are handled explicitly:
      footer      rejected - it is a total, never a record
      group       blank cells mean "same as the row above", so values carry down
      sub-detail  label/value pairs that belong to the record above, merged into it
    """
    # The banner band also has label/value rows with three or more cells
    # ("Packaged Facility Name | Twisted Growers LLC | Lab Facility Name"), so
    # "first row with three cells" picks the banner, not the header. The column
    # header is the WIDEST row in the opening band - it names every column.
    head_at, widest = None, 0
    for i, row in enumerate(grid[:30]):
        cells = [c for c in row if c]
        if len(cells) < 3 or all(re.fullmatch(r'[\d.,]+', c) for c in cells):
            continue
        # A sub-detail row can be WIDER than the header - Plants Destroyed prints
        # six label/value cells under a five-column header in some exports - so
        # the widest row is not always the header. A row that opens with a known
        # sub-detail label is never the header.
        if cells[0].lower().strip() in SUBDETAIL_LABELS:
            continue
        if len(cells) > widest:
            head_at, widest = i, len(cells)
    if head_at is None:
        return None, None

    head, seen = [], {}
    for c in grid[head_at]:
        c = re.sub(r'\s+', ' ', c).strip()
        if not c:
            head.append('')
            continue
        seen[c] = seen.get(c, 0) + 1
        head.append(c if seen[c] == 1 else '%s (%d)' % (c, seen[c]))
    width = sum(1 for h in head if h)

    carry = carry or ['Manifest', 'Inv. Nbr', 'Origin Lic.', 'Origin Facility', 'Origin Type',
                      'Dest. Lic.', 'Destination Facility', 'Dest. Type', 'Type',
                      'Created', 'Received', 'Voided',
                      'Harvest Batch', 'Strain', 'Created Date', 'Lab Testing', 'Finished']

    rows, last = [], {}
    for r in grid[head_at + 1:]:
        cells = [c for c in r if c]
        if not cells:
            continue
        if any(FOOTER.fullmatch(c) for c in cells):
            continue                                    # grand total or subtotal band

        # Sub-detail band: label/value pairs continuing the record above.
        # Identified by the LABEL alone. An earlier version also required the row
        # to be narrower than 60% of the header. That held on the real Plants
        # Destroyed file, whose header is 18 columns wide, and silently stopped
        # working on a narrower one - a rule that depends on the header's width
        # is not a rule about sub-detail rows. Found by a fixture, not in
        # production, which is the entire point of having fixtures.
        if rows and cells[0].lower().strip() in SUBDETAIL_LABELS:
            for i in range(0, len(cells) - 1, 2):
                label, value = cells[i].strip(), cells[i + 1]
                if label.lower() in SUBDETAIL_LABELS:
                    rows[-1][label] = value
            continue

        o = {k: (r[i] if i < len(r) else '') for i, k in enumerate(head) if k}
        for k in carry:                                  # group band carries down
            if k in o:
                if o[k]:
                    last[k] = o[k]
                elif k in last:
                    o[k] = last[k]
        rows.append(o)
    return head, rows


LICENCE_COLS = ['Origin Lic.', 'OriginLicenseNumber', 'Origin License',
                'Dest. Lic.', 'DestinationLicenseNumber', 'Destination License',
                'Packaged Lic. No.', 'Packaged Facility License', 'Licence', 'License']


def row_licence(row, fallback=None):
    """Which of OUR licences owns this row.

    A grid export can span BOTH licences. TransfersReport.csv holds 13,553
    manufacturing rows and 3,440 cultivation rows in one file; stamping a single
    licence on it would misfile 3,440 rows - the same fault that once stamped
    13,294 rows with a customer's retail licence.

    So the licence is read per row, from whichever column actually names a
    licence of OURS. On an outbound transfer that is the origin; on an inbound
    one the origin is the supplier and the destination is us. Taking the first
    column that holds one of our own licences handles both without needing to
    know which direction the report faces. A row that names neither falls back
    to the file-level licence, which comes from the filename.
    """
    for c in LICENCE_COLS:
        v = (row.get(c) or '').strip()
        if v in OUR_LICENCES:
            return v
    return fallback


def split_by_licence(rows, fallback=None):
    """Group rows by the licence that owns them, preserving order."""
    groups = {}
    for r in rows:
        groups.setdefault(row_licence(r, fallback), []).append(r)
    return groups


def stated_total(grid):
    """Metrc prints "Total Records: N" above every report. An import that lands a
    different number has lost or invented rows, so this is passed through and the
    database refuses to report success on a mismatch."""
    for row in grid[:20]:
        for c in row:
            m = re.search(r'Total Records:\s*([\d,]+)', str(c))
            if m:
                return int(m.group(1).replace(',', ''))
    return None


def banner(grid):
    """What Metrc prints above the table: the report title, the date it was run
    for, and the licence. Point in Time puts its as-of date on the line directly
    under the title - there is never a need to ask a person for it."""
    title = as_of = licence = None
    for i, row in enumerate(grid[:14]):
        cells = [c.strip() for c in row if c and c.strip()]
        if not cells:
            continue
        if title is None and len(cells) == 1 and not re.fullmatch(r'[\d/.,\- ]+', cells[0]):
            title = cells[0]
            continue
        if as_of is None and len(cells) == 1:
            m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', cells[0])
            if m:
                as_of = '%s-%02d-%02d' % (m.group(3), int(m.group(1)), int(m.group(2)))
                continue
            m = re.search(r'To\s+(\d{1,2})/(\d{1,2})/(\d{4})', cells[0])
            if m:
                as_of = '%s-%02d-%02d' % (m.group(3), int(m.group(1)), int(m.group(2)))
                continue
        # Only our own licences. A grid export has no banner, so scanning its
        # opening rows finds the DESTINATION licence in the first data row - which
        # once stamped 13,294 of our manufacturing rows with a customer's retail
        # licence. A licence on our own records can only ever be one of ours.
        if licence is None:
            for c in cells:
                if c in OUR_LICENCES:
                    licence = c
                    break
    return title, as_of, licence


def excel_date(v):
    """Metrc writes dates as Excel serials in some exports."""
    try:
        n = float(str(v).replace(',', ''))
    except Exception:
        return None
    if 20000 <= n <= 60000:
        import datetime
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(n))).isoformat()
    return None


def as_of_from(rows, head):
    """Latest event date inside the file. A report cannot know about tomorrow."""
    best = None
    fields = [f for f in DATE_FIELDS if f in head]
    for r in rows:
        for f in fields:
            v = r.get(f, '')
            iso = excel_date(v)
            if not iso:
                m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', str(v))
                if m:
                    iso = '%s-%02d-%02d' % (m.group(3), int(m.group(1)), int(m.group(2)))
                elif re.match(r'^\d{4}-\d{2}-\d{2}', str(v)):
                    iso = str(v)[:10]
            if iso and (best is None or iso > best):
                best = iso
    return best


def post(rows, licence, as_of, name, stated=None):
    """Send the file, chunking when it is too large for one request.

    The row-count check is made on the TOTAL stored across every chunk. Telling
    each chunk the whole file's count would make every chunk look short.
    """
    out, step = [], 4000
    chunked = len(rows) > step
    for i in range(0, len(rows), step):
        body = json.dumps({
            'rows': rows[i:i + step], 'licence': licence, 'as_of': as_of,
            'file_name': name,
            'stated_total': None if chunked else stated}).encode()
        # Long uploads drop occasionally ("EOF occurred in violation of protocol").
        # That is the connection, not the data, and a retry succeeds - but only a
        # TRANSPORT failure is retried. An error the server actually returned is
        # its answer and must not be papered over by trying again.
        for attempt in range(4):
            try:
                req = urllib.request.Request(FN, data=body, headers=HDR, method='POST')
                with urllib.request.urlopen(req, timeout=900) as r:
                    out.append(json.loads(r.read()))
                break
            except urllib.error.HTTPError:
                raise
            except Exception as e:
                if attempt == 3:
                    raise
                wait = 5 * (attempt + 1)
                print('    chunk %d: %s - retrying in %ds (%d of 3)'
                      % (i // step + 1, str(e)[:48], wait, attempt + 1))
                time.sleep(wait)

    if chunked and stated:
        ids = [o.get('result', {}).get('import_id') for o in out]
        ids = [i for i in ids if i]
        stored = sum(o.get('result', {}).get('rows_stored', 0) for o in out)
        v = urllib.request.Request(FN, data=json.dumps({
            'verify': {'imports': ids, 'stated': stated, 'stored': stored}}).encode(),
            headers=HDR, method='POST')
        with urllib.request.urlopen(v, timeout=300) as r:
            verdict = (json.loads(r.read()) or {}).get('result', {})
        out.append({'result': {'ok': verdict.get('ok'), 'rows_stored': 0,
                               'count_check': verdict.get('count_check'),
                               'report': out[0].get('result', {}).get('report'),
                               'title': out[0].get('result', {}).get('title')}})
    return out


def run(paths):
  """Import each file. Kept out of module scope so that importing this module
  for its parsing functions - as the fixture suite does - does not fire off an
  import of every file in the Downloads folder."""
  for path in paths:
    name = os.path.basename(path)
    try:
        if path.lower().endswith('.xlsx'):
            grid = grid_xlsx(path)
        elif path.lower().endswith('.xls'):
            grid = grid_xls(path)
        else:
            import csv
            grid = list(csv.reader(open(path, encoding='utf-8-sig', errors='ignore')))
        head, rows = to_rows(grid)
        if not rows:
            print('%-56s SKIP  no header found' % name[:56])
            continue
        # The FILENAME is stated; anything scanned out of the sheet is inferred.
        # A grid export has no banner, so scanning its opening rows finds the
        # destination licence in the first data row - which once stamped 13,294
        # manufacturing rows with a customer's retail licence. Filename wins, and
        # the result must be one of ours or it is discarded.
        title, banner_as_of, banner_lic = banner(grid)
        m = re.search(r'\bM[CP]\d{6}\b', name)
        lic = (m.group(0) if m else None) or banner_lic
        if lic not in OUR_LICENCES:
            lic = None
        as_of = banner_as_of or as_of_from(rows, head)
        stated = stated_total(grid)

        # A file may span both licences. Post each licence's rows under its own
        # stamp rather than forcing one licence onto the whole file.
        groups = split_by_licence(rows, lic)
        multi = len(groups) > 1
        total_stored, all_ok, report_name, notes = 0, True, None, []
        for glic, grows in sorted(groups.items(), key=lambda kv: (kv[0] or '')):
            # The stated total counts the whole file, so it can only be checked
            # against a single-group post. Split files are checked on the sum.
            res = post(grows, glic, as_of, name, None if multi else stated)
            first = res[0].get('result', {})
            if len(res) > 1:
                first = {**first, 'ok': res[-1].get('result', {}).get('ok', first.get('ok'))}
            stored = sum(r.get('result', {}).get('rows_stored', 0) for r in res)
            total_stored += stored
            report_name = report_name or first.get('report')
            all_ok = all_ok and bool(first.get('ok'))
            if not first.get('ok'):
                notes.append('%s FAILED %s' % (glic, str(first.get('error'))[:60]))
            elif multi:
                notes.append('%s=%d' % (glic or 'unattributed', stored))
            chk = (res[-1].get('result', {}) or {}).get('count_check') or first.get('count_check') or ''

        if all_ok:
            verdict = 'OK' if (stated is None or total_stored == stated) else \
                      'COUNT MISMATCH stated=%s stored=%s' % (stated, total_stored)
            print('%-52s -> %-22s stated=%-6s stored=%-6s %s %s' %
                  (name[:52], report_name, stated, total_stored, verdict,
                   ('[' + ' '.join(notes) + ']') if multi else ''))
        else:
            print('%-52s FAILED  %s' % (name[:52], '; '.join(notes)[:90]))
    except Exception as e:
        print('%-56s ERROR %s' % (name[:56], str(e)[:80]))


def main():
    if not os.environ.get('SUPA_KEY'):
        print('SUPA_KEY is not set. Importing needs it; parsing does not.')
        return 2
    targets = sys.argv[1:] or (
        glob.glob(r'C:\Users\demar\Downloads\Metrc-Massachusetts-*.xlsx') +
        glob.glob(r'C:\Users\demar\Downloads\*Report*.xls') +
        glob.glob(r'C:\Users\demar\Downloads\*Report*.csv'))
    run(sorted(set(targets)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
